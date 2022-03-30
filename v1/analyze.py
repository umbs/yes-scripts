import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Color
from openpyxl.chart import (
        PieChart,
        Reference, Series, PieChart3D
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
import win32com.client
import PIL
from PIL import ImageGrab, Image
import os
import sys
import docx
from docx.shared import Inches
import subprocess

MAX_ROWS=400
CURRENT_ROW=1
HEIGHT=8
WIDTH=11
wb = openpyxl.load_workbook("./responses.xlsx")
res_wb = openpyxl.Workbook()
name_of_sheet = ""
inputFolderPath = os.getcwd()
inputExcelName = "summary.xlsx"
inputExcelFilePath = os.path.join(inputFolderPath, inputExcelName)
reportList = ["Hubbard", "Dorsa", "Lyndale", "Ben Painter", "Horace Cureton", "Linda Vista", "Renaissance"]

def get_sheet1():
    global inputExcelFilePath
        
    # Open the excel application using win32com
    o = win32com.client.Dispatch("Excel.Application")
    # Disable alerts and visibility to the user
    o.Visible = 0
    o.DisplayAlerts = False
    
    # Open workbook
    wx = o.Workbooks.Open(inputExcelFilePath)
    sheet = o.Sheets('Sheet')
    wx.Sheets(1).Delete()
    wx.Close(SaveChanges=True)
    o.Quit()


def get_pie_data(res_sheet):
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=1, max_row=5), Reference(res_sheet, min_col=3, min_row=1, max_row=5), "What did you like best about SKY Schools?", "E1", 1)
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=8, max_row=11), Reference(res_sheet, min_col=3, min_row=8, max_row=11), "Do you use what you have learned in SKY Schools?", "L1", 1)
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=14, max_row=16), Reference(res_sheet, min_col=3, min_row=14, max_row=16), " After SKY Schools do you feel: [More focused]", "S1", 1)
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=19, max_row=21), Reference(res_sheet, min_col=3, min_row=19, max_row=21), " After SKY Schools do you feel: [More calm and relaxed]", "E17", 1)
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=24, max_row=26), Reference(res_sheet, min_col=3, min_row=24, max_row=26), " After SKY Schools do you feel: [Happy]", "L17", 1)
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=29, max_row=31), Reference(res_sheet, min_col=3, min_row=29, max_row=31), " After SKY Schools do you feel: [Healthy]", "S17", 1)
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=34, max_row=36), Reference(res_sheet, min_col=3, min_row=34, max_row=36), " After SKY Schools do you feel: [Less stress]", "E33", 1)
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=39, max_row=40), Reference(res_sheet, min_col=3, min_row=39, max_row=40), "SKY Schools was [Fun]", "L33", 2)
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=43, max_row=44), Reference(res_sheet, min_col=3, min_row=43, max_row=44), "SKY Schools was [Interesting]", "S33", 2)
    draw_pie_chart(res_sheet, Reference(res_sheet, min_col=2, min_row=47, max_row=51), Reference(res_sheet, min_col=3, min_row=47, max_row=51), "What was your level of participation in SKY Schools?", "E49", 1)

    
def draw_pie_chart(res_sheet, pie_label, pie_data, pie_title, pie_place, pie_count):
    pie = PieChart()
    labels = pie_label
    data = pie_data
    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = pie_title
    pie.height = HEIGHT
    pie.width = WIDTH
    # Showing data labels as percentage
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    series = pie.series[0]
    #first series
    pt = openpyxl.chart.marker.DataPoint(idx=0)
    pt.graphicalProperties.solidFill = "00FF00"
    series.dPt.append(pt)
    # second series
    pt1 = openpyxl.chart.marker.DataPoint(idx=1)
    if pie_count == 1:
        pt1.graphicalProperties.solidFill = "00FFFF"
    elif pie_count == 2:
        pt1.graphicalProperties.solidFill = "FF0000"
    series.dPt.append(pt1)
    # third series
    pt2 = openpyxl.chart.marker.DataPoint(idx=2)
    pt2.graphicalProperties.solidFill = "FF0000"
    series.dPt.append(pt2)
    # fourth series
    pt3 = openpyxl.chart.marker.DataPoint(idx=3)
    pt3.graphicalProperties.solidFill = "FFFF00"
    series.dPt.append(pt3)
    # fifth series
    pt4 = openpyxl.chart.marker.DataPoint(idx=4)
    pt4.graphicalProperties.solidFill = "808080"
    series.dPt.append(pt4)
    res_sheet.add_chart(pie, pie_place)
    

def write_result(res_sheet, query, data):
    # Write data to results excel
    # Query name goes in first column A and results goes in B and C
    global CURRENT_ROW

    # Merge cells
    sz = len(data)
    res_sheet.merge_cells('A'+str(CURRENT_ROW)+':A'+str(CURRENT_ROW+sz))

    # Merged cell has the query, adjacent cells has responses
    res_sheet['A' + str(CURRENT_ROW)] = query

    # Center align
    res_sheet['A' + str(CURRENT_ROW)].alignment = Alignment(vertical='center', horizontal='center')

    total_count = 0

    # Responses
    for k, v in data.items():
        res_sheet['B' + str(CURRENT_ROW)] = k
        res_sheet['C' + str(CURRENT_ROW)] = v
        total_count += int(v)
        CURRENT_ROW += 1

    # res_sheet['B' + str(CURRENT_ROW)] = 'Total'
    # res_sheet['C' + str(CURRENT_ROW)] = total_count

    CURRENT_ROW += 2


def best_about_sky_schools(sheet, res_sheet):
    global CURRENT_ROW
    col = column_index_from_string('D')
    query = sheet.cell(row=1, column=col).value

    # Each row can have multiple answers, such as [Breathing, Yoga, All of it]
    # Each answer must be counted.
    data = dict()
    for i in range(2, MAX_ROWS):
        val = sheet.cell(row=i, column=col).value

        if val is None:
            # no more entries?
            break

        val = val.split(',')

        for v in val:
            v = v.strip()
            data[v] = 1 + data.get(v, 0)

    write_result(res_sheet, query, data)


def use_learning_from_sky_schools(sheet, res_sheet):
    global CURRENT_ROW
    col = column_index_from_string('E')
    query = sheet.cell(row=1, column=col).value

    # Each row can have 'Sometimes', 'Every day', 'Never' or some sentence which
    # is counted as 'Other'
    data = dict()
    for i in range(2, MAX_ROWS):
        val = sheet.cell(row=i, column=col).value

        if val is None:
            # no more entries?
            break

        val = val.strip()

        if val in ['Sometimes', 'Everyday', 'Never']:
            data[val] = 1 + data.get(val, 0)
        else:
            data['Other'] = 1 + data.get('Other', 0)

    write_result(res_sheet, query, data)


def how_do_you_feel(sheet, res_sheet):
    global CURRENT_ROW
    # There are 8 questions starting from column 'F'. Used a for loop to iterate
    # through them. Responses follow same format.
    for c in ['F', 'G', 'H', 'I', 'J']:
        col = column_index_from_string(c)
        query = sheet.cell(row=1, column=col).value

        # Each row can only have 'Yes', 'No', 'A little bit'
        data = dict()
        for i in range(2, MAX_ROWS):
            val = sheet.cell(row=i, column=col).value

            if val is None:
                # no more entries?
                break

            val = val.strip()

            data[val] = 1 + data.get(val, 0)

        write_result(res_sheet, query, data)


def sky_schools_was(sheet, res_sheet):
    global CURRENT_ROW
    # There are 3 questions starting from column 'M'. Used a for loop to iterate
    # through them. Responses follow same format.
    for c in ['K', 'L']:
        col = column_index_from_string(c)
        query = sheet.cell(row=1, column=col).value

        # Each row can only have 'Yes', 'No'
        data = dict()
        for i in range(2, MAX_ROWS):
            val = sheet.cell(row=i, column=col).value

            if val is None:
                # no more entries?
                break

            val = val.strip()

            data[val] = 1 + data.get(val, 0)

        write_result(res_sheet, query, data)


def sky_part(sheet, res_sheet):
    global CURRENT_ROW
    col = column_index_from_string('O')
    query = sheet.cell(row=1, column=col).value

    # Each row can have multiple answers, such as [Breathing, Yoga, All of it]
    # Each answer must be counted.
    data = dict()
    for i in range(2, MAX_ROWS):
        val = sheet.cell(row=i, column=col).value

        if type(val) == int or type(val) == float:
            str1 = str(int(val * 100))
            str2 = "%"
            val = str1+str2

        if val is None:
            # no more entries?
            break

        val = val.split(',')

        for v in val:
            v = v.strip()
            data[v] = 1 + data.get(v, 0)

    write_result(res_sheet, query, data)

def main():
    global CURRENT_ROW
    global reportList
    for name in wb.sheetnames:
        CURRENT_ROW = 1
        sheet = wb[name]
        res_sheet = res_wb.create_sheet(title=name)

        for i in range(len(reportList)):
            if sheet.title not in (reportList[i]):
                continue

        # For some reason max_row is coming out large value than actual entries
        # max_row = sheet.max_row

        # 1) What did you like best about SKY schools?
        best_about_sky_schools(sheet, res_sheet)

        # 2) Do you use what you learned in SKY schools?
        use_learning_from_sky_schools(sheet, res_sheet)

        # 3) After SKY Schools do you feel: More focused, More calm and so on?
        how_do_you_feel(sheet, res_sheet)

        # 4) SKY schools was fun/interesting/relaxing
        sky_schools_was(sheet, res_sheet)

        # 5) What was your level of participation in SKY Schools?
        sky_part(sheet, res_sheet)

        # Draw pie charts with results
        get_pie_data(res_sheet)

        res_wb.save('summary.xlsx')
        
        get_sheet1()
    


if __name__ == "__main__":
    main()
