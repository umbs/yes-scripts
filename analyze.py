import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.chart import (
        PieChart,
        Reference
)

MAX_ROWS=400
CURRENT_ROW=1

wb = openpyxl.load_workbook("./responses.xlsx")
res_wb = openpyxl.Workbook()


def best_about_sky_schools(sheet, res_sheet):
    col = column_index_from_string('D')
    query = sheet.cell(row=1, column=col).value

    # Each row can have multiple answers, such as [Breathing, Yoga, All of it]
    # Each answer must be counted.
    # Atmost 5 possible answers: 
    #   'Games', 'Breathing', 'Yoga', 'The Golden # Keys', 'All of it'.
    total_votes = 0
    data = dict()
    for i in range(2, MAX_ROWS):
        val = sheet.cell(row=i, column=col).value

        if val is None:
            # no more entries?
            break

        val = val.split(',')

        for v in val:
            v = v.strip()
            data[v] = 1 + data.get(v, 0)
            total_votes += 1

    # print(data, total_votes)

    # Write data to results excel
    # Query name goes in first column (A) and rows 1-5
    # Results go in columns B and C and rows 1-5
    res_sheet.merge_cells('A1:A5')
    res_sheet['A1'] = query
    dest_row = 1
    for k, v in data.items():
        res_sheet['B' + str(dest_row)] = k
        res_sheet['C' + str(dest_row)] = v
        dest_row += 1


def use_learning_from_sky_schools(sheet, res_sheet):
    col = column_index_from_string('E')
    query = sheet.cell(row=1, column=col).value

    # Each row can have 'Sometimes', 'Every day', 'Never' or some sentence which
    # is counted as 'Other'
    total_votes = 0
    data = dict()
    for i in range(2, MAX_ROWS):
        val = sheet.cell(row=i, column=col).value

        if val is None:
            # no more entries?
            break

        val = val.strip()

        if val in ['Sometimes', 'Every day', 'Never']:
            data[val] = 1 + data.get(val, 0)
        else:
            data['Other'] = 1 + data.get('Other', 0)

        total_votes += 1

    # print(data, total_votes)

    # Write data to results excel
    # Query name goes in first column (A) and rows 7-10
    res_sheet['A7'] = query
    dest_row = 2
    for k, v in data.items():
        res_sheet['B' + str(dest_row)] = k
        res_sheet['C' + str(dest_row)] = v
        dest_row += 1


def how_do_you_feel(sheet, res_sheet):
    # There are 8 questions starting from column 'F'. Used a for loop to iterate
    # through them. Responses follow same format.
    dest_col = 7
    for c in ['F', 'G', 'H', 'I', 'J', 'K', 'L']:
        col = column_index_from_string(c)
        query = sheet.cell(row=1, column=col).value

        # Each row can only have 'Yes', 'No', 'A little bit'
        total_votes = 0
        data = dict()
        for i in range(2, MAX_ROWS):
            val = sheet.cell(row=i, column=col).value

            if val is None:
                # no more entries?
                break

            val = val.strip()

            data[val] = 1 + data.get(val, 0)
            total_votes += 1

        # Write data to results excel
        # Query name goes in first cell
        dest_col_letter = get_column_letter(dest_col)
        dest_row = 2
        res_sheet[dest_col_letter + '1'] = query

        for k, v in data.items():
            res_sheet[dest_col_letter + str(dest_row)] = k
            res_sheet[get_column_letter(dest_col+1) + str(dest_row)] = v
            dest_row += 1

        dest_col += 3


def draw_charts(res_sheet):
    pass


def main():
    for name in wb.sheetnames:
        sheet = wb[name]
        res_sheet = res_wb.create_sheet(title=name)

        if sheet.title != 'Linda Vista ES':
            continue

        # For some reason max_row is coming out large value than actual entries
        # max_row = sheet.max_row

        # 1) What did you like best about SKY schools?
        best_about_sky_schools(sheet, res_sheet)

        # 2) Do you use what you learned in SKY schools?
        use_learning_from_sky_schools(sheet, res_sheet)

        # 3) After SKY Schools do you feel: More focused, More calm and so on?
        how_do_you_feel(sheet, res_sheet)

        res_wb.save('summary.xlsx')


if __name__ == "__main__":
    main()
